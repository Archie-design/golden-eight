#!/usr/bin/env python3
"""
匯入 Google Form 歷史打卡資料到 checkin_records 表。

讀取 .env.local 取得 Supabase 憑證。

用法：
  python3 scripts/import_form_data.py            # 正式匯入
  python3 scripts/import_form_data.py --dry-run  # 僅預覽，不寫入 DB
"""

import json
import os
import sys
import urllib.request
import urllib.error
from collections import defaultdict
from datetime import timezone, timedelta
from pathlib import Path

# ---------- 設定 ----------
EXCEL_PATH = Path(__file__).parent.parent / 'reference' / '星光🌟黃金八套餐 登記表 (回覆).xlsx'
ENV_PATH   = Path(__file__).parent.parent / '.env.local'
BATCH_SIZE = 50

# 台北時區 UTC+8
TZ_TAIPEI = timezone(timedelta(hours=8))

# 名稱正規化規則
NAME_NORMALIZE = {
    '昀洋':        '戴昀洋',
    '黃名禎（波波）': '黃名禎',
    '思全':        '賴思全',
}

# 表單任務文字 → tasks 陣列索引
TASK_MAP = {
    '早睡早起':         0,
    '破曉打拳':         1,
    '丹氣跑步':         2,
    '曬太陽':           3,
    '工作8小時':        4,
    '不吃肉':           5,
    '寫觀心書/覺察日記': 6,
    '做淨心功法':       7,
}

# ---------- 環境變數讀取 ----------

def load_env(path: Path) -> dict:
    env = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        key, _, val = line.partition('=')
        env[key.strip()] = val.strip().strip('"').strip("'")
    return env

# ---------- Supabase REST 封裝 ----------

class SupabaseClient:
    def __init__(self, url: str, key: str):
        self.url = url.rstrip('/')
        self.headers = {
            'apikey':        key,
            'Authorization': f'Bearer {key}',
            'Content-Type':  'application/json',
            'Prefer':        'return=minimal',
        }

    def _request(self, method: str, path: str, body=None):
        full_url = f'{self.url}/rest/v1{path}'
        data = json.dumps(body).encode() if body is not None else None
        req = urllib.request.Request(full_url, data=data, headers=self.headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            msg = e.read().decode(errors='replace')
            raise RuntimeError(f'HTTP {e.code} {method} {path}: {msg}') from e

    def get(self, path: str) -> list:
        return self._request('GET', path)

    def post(self, path: str, body: list) -> None:
        self._request('POST', path, body)

# ---------- 資料解析 ----------

def parse_excel(path: Path) -> dict:
    """
    回傳 {(normalized_name, date_str): record_dict}
    每個 (name, date) 只保留時間戳記最晚的那筆。
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit('❌ 缺少 openpyxl，請先執行：pip install openpyxl')

    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    groups: dict = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        ts         = ws.cell(row, 1).value
        name_raw   = ws.cell(row, 2).value
        tasks_str  = ws.cell(row, 3).value
        if not ts or not name_raw:
            continue

        name = NAME_NORMALIZE.get(name_raw.strip(), name_raw.strip())
        ts_taipei = ts.replace(tzinfo=TZ_TAIPEI)
        if ts_taipei.hour < 12:
            date_str = (ts_taipei.date() - timedelta(days=1)).isoformat()
        else:
            date_str = ts_taipei.date().isoformat()

        # 解析任務 boolean[8]
        tasks = [False] * 8
        if tasks_str:
            for t in tasks_str.split(','):
                idx = TASK_MAP.get(t.strip())
                if idx is not None:
                    tasks[idx] = True

        base_score  = sum(tasks)
        total_score = base_score

        # submit_time：假設表單時間為台北時間，轉為 ISO 8601 UTC offset
        submit_time = ts.replace(tzinfo=TZ_TAIPEI).isoformat()

        groups[(name, date_str)].append({
            'ts':           ts,
            'tasks':        tasks,
            'base_score':   base_score,
            'total_score':  total_score,
            'submit_time':  submit_time,
        })

    # 每組取最新一筆
    result = {}
    for (name, date_str), entries in groups.items():
        latest = max(entries, key=lambda x: x['ts'])
        result[(name, date_str)] = latest
    return result

# ---------- 主程式 ----------

def main():
    dry_run = '--dry-run' in sys.argv
    if dry_run:
        print('=== DRY-RUN 模式（不寫入 DB）===\n')

    # 讀取憑證
    env = load_env(ENV_PATH)
    supabase_url = (
        os.environ.get('NEXT_PUBLIC_SUPABASE_URL')
        or env.get('NEXT_PUBLIC_SUPABASE_URL')
    )
    service_key = (
        os.environ.get('SUPABASE_SERVICE_ROLE_KEY')
        or env.get('SUPABASE_SERVICE_ROLE_KEY')
    )
    if not supabase_url or not service_key:
        sys.exit('❌ 找不到 NEXT_PUBLIC_SUPABASE_URL 或 SUPABASE_SERVICE_ROLE_KEY，請確認 .env.local')

    db = SupabaseClient(supabase_url, service_key)

    # Step 1：解析 Excel
    print('📂 讀取 Excel…')
    form_records = parse_excel(EXCEL_PATH)
    print(f'   去重後記錄數：{len(form_records)}')

    # Step 2：查詢 DB 成員清單
    print('\n👥 查詢成員清單…')
    members_raw = db.get('/members?select=id,name&order=id')
    name_to_id: dict[str, str] = {m['name']: m['id'] for m in members_raw}
    print(f'   DB 成員數：{len(name_to_id)}')

    # 找出無法對應的姓名
    form_names = {name for (name, _) in form_records}
    unmatched = sorted(form_names - set(name_to_id))
    if unmatched:
        print(f'\n⚠️  以下表單姓名在 DB 中找不到對應成員（共 {len(unmatched)} 人，將略過）：')
        for n in unmatched:
            count = sum(1 for (nm, _) in form_records if nm == n)
            print(f'   - {n}（{count} 筆）')
    else:
        print('   ✅ 所有表單姓名均已對應到 DB 成員')

    # Step 3：查詢既有記錄
    print('\n🗄️  查詢既有打卡記錄…')
    existing_raw = db.get('/checkin_records?select=member_id,date&date=gte.2026-04-01')
    existing: set[tuple] = {(r['member_id'], r['date']) for r in existing_raw}
    print(f'   既有記錄數（2026-04 起）：{len(existing)}')

    # Step 4：篩選需匯入的記錄
    to_insert = []
    skip_no_member = 0
    skip_conflict  = 0

    for (name, date_str), rec in sorted(form_records.items()):
        member_id = name_to_id.get(name)
        if not member_id:
            skip_no_member += 1
            continue
        if (member_id, date_str) in existing:
            skip_conflict += 1
            continue
        to_insert.append({
            'member_id':   member_id,
            'date':        date_str,
            'tasks':       rec['tasks'],
            'base_score':  rec['base_score'],
            'punch_bonus': 0,
            'total_score': rec['total_score'],
            'punch_streak': 0,
            'work_hours':  None,
            'note':        None,
            'submit_time': rec['submit_time'],
        })

    print(f'\n📊 匯入摘要')
    print(f'   名稱無法對應（略過）：{skip_no_member} 筆')
    print(f'   已有 DB 記錄（略過）：{skip_conflict} 筆')
    print(f'   待匯入：            {len(to_insert)} 筆')

    if not to_insert:
        print('\n✅ 無需匯入任何記錄。')
        return

    # 印出 dry-run 預覽
    if dry_run:
        print('\n--- DRY-RUN 預覽（前 10 筆）---')
        for rec in to_insert[:10]:
            score = rec['total_score']
            tasks_str = ''.join('✓' if t else '·' for t in rec['tasks'])
            print(f"  {rec['member_id']} {rec['date']} [{tasks_str}] score={score}")
        if len(to_insert) > 10:
            print(f'  … 共 {len(to_insert)} 筆')
        print('\n✅ Dry-run 完成，未寫入任何資料。')
        return

    # Step 5：批次寫入
    print(f'\n⬆️  開始寫入（批次大小 {BATCH_SIZE}）…')
    inserted = 0
    for i in range(0, len(to_insert), BATCH_SIZE):
        batch = to_insert[i:i + BATCH_SIZE]
        db.post('/checkin_records', batch)
        inserted += len(batch)
        print(f'   {inserted}/{len(to_insert)}', end='\r', flush=True)

    print(f'\n\n✅ 完成！成功匯入 {inserted} 筆打卡記錄。')


if __name__ == '__main__':
    main()
