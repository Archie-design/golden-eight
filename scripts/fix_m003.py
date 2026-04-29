#!/usr/bin/env python3
"""
比對並修正黃琳貽 (M003) 的打卡紀錄。
以 Excel 表單資料為標準，套用中午換日規則，更新資料庫。
"""

import json
import os
import sys
import urllib.request
import urllib.error
from collections import defaultdict
from datetime import timezone, timedelta
from pathlib import Path

EXCEL_PATH = Path(__file__).parent.parent / 'reference' / '星光🌟黃金八套餐 登記表 (回覆).xlsx'
ENV_PATH   = Path(__file__).parent.parent / '.env.local'
MEMBER_ID  = 'M003'
TARGET_NAME = '黃琳貽'

TZ_TAIPEI = timezone(timedelta(hours=8))

TASK_MAP = {
    '早睡早起':         0,
    '破曉打拳':         1,
    '丹氣跑步':         2,
    '曬太陽':           3,
    '工作8小時':        4,
    '不吃肉':           5,
    '寫觀心書/覺察日記': 6,
    '做淨心功法':       7,
}

def load_env(path):
    env = {}
    if not path.exists():
        return env
    for line in path.read_text(encoding='utf-8').splitlines():
        line = line.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        k, _, v = line.partition('=')
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env

class SupabaseClient:
    def __init__(self, url, key):
        self.url = url.rstrip('/')
        self.base_headers = {
            'apikey': key,
            'Authorization': f'Bearer {key}',
            'Content-Type': 'application/json',
            'Prefer': 'return=minimal',
        }

    def get(self, path):
        req = urllib.request.Request(
            f'{self.url}/rest/v1{path}',
            headers=self.base_headers,
        )
        with urllib.request.urlopen(req) as r:
            return json.loads(r.read())

    def patch(self, path, body):
        data = json.dumps(body).encode()
        req = urllib.request.Request(
            f'{self.url}/rest/v1{path}',
            data=data,
            headers=self.base_headers,
            method='PATCH',
        )
        with urllib.request.urlopen(req) as r:
            pass

    def post(self, path, body):
        data = json.dumps(body).encode()
        req = urllib.request.Request(
            f'{self.url}/rest/v1{path}',
            data=data,
            headers=self.base_headers,
            method='POST',
        )
        with urllib.request.urlopen(req) as r:
            pass

def parse_excel_for_member(path, target_name):
    try:
        import openpyxl
    except ImportError:
        sys.exit('❌ 缺少 openpyxl，請先執行：pip install openpyxl')

    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    groups = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        ts        = ws.cell(row, 1).value
        name_raw  = ws.cell(row, 2).value
        tasks_str = ws.cell(row, 3).value
        if not ts or not name_raw:
            continue
        if name_raw.strip() != target_name:
            continue

        ts_taipei = ts.replace(tzinfo=TZ_TAIPEI)
        if ts_taipei.hour < 12:
            date_str = (ts_taipei.date() - timedelta(days=1)).isoformat()
        else:
            date_str = ts_taipei.date().isoformat()

        tasks = [False] * 8
        if tasks_str:
            for t in tasks_str.split(','):
                idx = TASK_MAP.get(t.strip())
                if idx is not None:
                    tasks[idx] = True

        base_score   = sum(tasks)
        submit_time  = ts_taipei.isoformat()

        groups[date_str].append({
            'ts':          ts,
            'tasks':       tasks,
            'base_score':  base_score,
            'submit_time': submit_time,
        })

    result = {}
    for date_str, entries in groups.items():
        latest = max(entries, key=lambda x: x['ts'])
        result[date_str] = latest
    return result

def main():
    dry_run = '--dry-run' in sys.argv

    env = load_env(ENV_PATH)
    supabase_url = os.environ.get('NEXT_PUBLIC_SUPABASE_URL') or env.get('NEXT_PUBLIC_SUPABASE_URL')
    service_key  = os.environ.get('SUPABASE_SERVICE_ROLE_KEY') or env.get('SUPABASE_SERVICE_ROLE_KEY')
    if not supabase_url or not service_key:
        sys.exit('❌ 找不到 Supabase 憑證')

    db = SupabaseClient(supabase_url, service_key)

    # 1. 從 Excel 讀取 M003 表單資料
    print(f'📂 讀取 Excel（{TARGET_NAME}）…')
    form_records = parse_excel_for_member(EXCEL_PATH, TARGET_NAME)
    print(f'   表單記錄（中午換日後）：{len(form_records)} 筆')

    # 2. 查詢 DB 現有記錄
    print('\n🗄️  查詢 DB 記錄…')
    db_recs_raw = db.get(f'/checkin_records?member_id=eq.{MEMBER_ID}&date=gte.2026-04-01&order=date&select=id,date,tasks,base_score,total_score,submit_time')
    db_by_date = {r['date']: r for r in db_recs_raw}
    print(f'   DB 四月記錄：{len(db_by_date)} 筆')

    # 3. 比對
    to_insert = []
    to_update = []  # (id, date, patch_body, reason)

    all_dates = sorted(set(form_records) | set(db_by_date))
    print('\n📊 比對結果：')
    for date in all_dates:
        form = form_records.get(date)
        db   = db_by_date.get(date)

        if form and not db:
            tasks_str = ''.join('✓' if t else '·' for t in form['tasks'])
            print(f'  ➕ {date}  [{tasks_str}]  score={form["base_score"]}  → 新增')
            to_insert.append({
                'member_id':   MEMBER_ID,
                'date':        date,
                'tasks':       form['tasks'],
                'base_score':  form['base_score'],
                'punch_bonus': 0,
                'total_score': float(form['base_score']),
                'punch_streak': 0,
                'work_hours':  None,
                'note':        None,
                'submit_time': form['submit_time'],
            })
        elif form and db:
            # 比較 tasks
            form_tasks_str = ''.join('✓' if t else '·' for t in form['tasks'])
            db_tasks_str   = ''.join('✓' if t else '·' for t in db['tasks'])
            if form['tasks'] != db['tasks'] or form['base_score'] != db['base_score']:
                print(f'  ✏️  {date}  DB=[{db_tasks_str}] score={db["base_score"]}  →  表單=[{form_tasks_str}] score={form["base_score"]}')
                to_update.append((db['id'], date, {
                    'tasks':       form['tasks'],
                    'base_score':  form['base_score'],
                    'total_score': float(form['base_score']),
                    'submit_time': form['submit_time'],
                }))
            else:
                print(f'  ✅ {date}  [{form_tasks_str}]  score={form["base_score"]}  一致')
        else:
            db_tasks_str = ''.join('✓' if t else '·' for t in db['tasks'])
            print(f'  ⚠️  {date}  [{db_tasks_str}]  score={db["base_score"]}  僅 DB 有（保留）')

    print(f'\n   需新增：{len(to_insert)} 筆　需更新：{len(to_update)} 筆')

    if dry_run:
        print('\n✅ Dry-run 完成，未寫入。')
        return

    if not to_insert and not to_update:
        print('\n✅ 無需修改。')
        return

    # 4. 寫入
    supabase_client = SupabaseClient(supabase_url, service_key)

    inserted = 0
    for row in to_insert:
        supabase_client.post('/checkin_records', [row])
        inserted += 1
    print(f'\n⬆️  新增 {inserted} 筆')

    updated = 0
    for rec_id, date, patch in to_update:
        supabase_client.patch(f'/checkin_records?id=eq.{rec_id}', patch)
        updated += 1
    print(f'✏️  更新 {updated} 筆')

    print('\n✅ 完成！')

if __name__ == '__main__':
    main()
